#!/usr/bin/env python3
"""
generate_excel.py — Génère un fichier Excel des échanges avec l'IA pour le BAS Silvercrest.

Livrable : deliverables/echanges_ia.xlsx

Usage :
    python deliverables/generate_excel.py
"""

import subprocess
import sys
import os
import re
import json
from pathlib import Path

# ─── Auto-install openpyxl ────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("⚙️  openpyxl non installé — installation en cours...")
    # Essai 1 : pip standard
    ret = subprocess.call([sys.executable, "-m", "pip", "install", "openpyxl"],
                          stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    # Essai 2 : Debian/Ubuntu --break-system-packages
    if ret != 0:
        ret = subprocess.call(
            [sys.executable, "-m", "pip", "install", "openpyxl",
             "--break-system-packages"],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
    # Essai 3 : pip utilisateur
    if ret != 0:
        ret = subprocess.call(
            [sys.executable, "-m", "pip", "install", "openpyxl", "--user"],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
    if ret != 0:
        print("❌  Impossible d'installer openpyxl automatiquement.")
        print("    Lancez manuellement : pip install openpyxl")
        sys.exit(1)
    print("✅  openpyxl installé.")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

# ─── Chemins ──────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent.parent   # …/sysml-agent/
RESULTS_DIR = BASE_DIR / "experiments" / "results" / "bleed_air_system"
OUT_FILE = BASE_DIR / "deliverables" / "echanges_ia.xlsx"
API_BASE = "http://localhost:8000"

# ─── Styles Excel ─────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
ROW_FILL_ALT = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
ROW_FILL_WHT = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
ALIGN_WRAP_TOP   = Alignment(wrap_text=True, vertical="top")
ALIGN_WRAP_CTR   = Alignment(wrap_text=True, vertical="center", horizontal="center")

# ─── Données fixes — Feuille "Résumé" ─────────────────────────────────────────
SUMMARY_HEADERS = [
    "Système",
    "Version",
    "Niveau",
    "Taille description (lignes)",
    "Taille SysML v2 (chars)",
    "Nb stakeholders",
    "Nb use cases",
    "Nb warnings",
    "Couverture vs référence",
]

SUMMARY_ROWS = [
    ("BAS Silvercrest", "V1", "Opérationnel",  56,  3166,  3,   4,  0,  "54%"),
    ("BAS Silvercrest", "V1", "Fonctionnel",   56,  4200,  "-", "-", "-", "53%"),
    ("BAS Silvercrest", "V1", "Logique",       56,  1059,  "-", "-", "-", "-"),
    ("BAS Silvercrest", "V1", "Technique",     56, 12290,  "-", "-", "-", "-"),
    ("BAS Silvercrest", "V2", "Opérationnel", 155,  4456,  5,   8, "-", "75%"),
    ("BAS Silvercrest", "V2", "Fonctionnel",  155,  1241,  "-", "-", "-", "60%"),
    ("BAS Silvercrest", "V2", "Logique",      155,  1204,  "-", "-", "-", "-"),
    ("BAS Silvercrest", "V2", "Technique",    155,  1032,  "-", "-", "-", "-"),
]

# ─── Commentaires automatiques ────────────────────────────────────────────────
COMMENTS: dict[tuple, str] = {
    ("V1", "Opérationnel"): "stakeholders=[], 4 UC génériques. Couverture 54%",
    ("V2", "Opérationnel"): "5 stakeholders, 8 UC décomposés. Couverture 75%. Corrections P5+P8 appliquées",
    ("V1", "Fonctionnel"):  "18 flux non typés. Couverture 53%",
    ("V2", "Fonctionnel"):  "Flux typés (P6). Notation composant::fonction (P7). Code condensé (1241 chars vs 4200)",
    ("V1", "Logique"):      "Composants et connexions basiques",
    ("V2", "Logique"):      "Légère amélioration (+14%)",
    ("V1", "Technique"):    "Code détaillé (12290 chars)",
    ("V2", "Technique"):    "Code condensé (1032 chars). Régression due au contexte trop long",
}

# ─── Feuille "Échanges détaillés" — en-têtes ─────────────────────────────────
DETAIL_HEADERS = [
    "#",
    "Système",
    "Version",
    "Niveau",
    "Prompt envoyé",
    "Réponse IA (JSON)",
    "Code SysML v2",
    "Commentaires d'analyse",
]

# ─── Données fixes — Feuille "Prompts types" ──────────────────────────────────
PROMPTS_HEADERS = [
    "ID",
    "Nom",
    "Niveau MBSE",
    "Problème résolu",
    "Règle ajoutée",
    "Impact mesuré",
]

PROMPTS_TYPES = [
    (
        "P1", "Stakeholder = personne", "Opérationnel",
        "Confusion stakeholder/système externe",
        "Un stakeholder est TOUJOURS une personne. Un équipement est un système externe.",
        "Stakeholders correctement classifiés",
    ),
    (
        "P2", "Exigences mesurables", "Opérationnel",
        "Sur-génération d'exigences (8 au lieu de 2)",
        "Les exigences sont UNIQUEMENT des contraintes mesurables avec des chiffres",
        "Exigences limitées aux valeurs chiffrées",
    ),
    (
        "P3", "Connexions valides", "Logique",
        "Connexions référençant des composants inexistants",
        'Toute connexion doit lier deux composants définis dans "parts"',
        "0 connexion invalide",
    ),
    (
        "P4", "Exigences allouées", "Logique",
        "0 exigences allouées aux composants",
        "Si des exigences existent, elles DOIVENT être allouées",
        "Exigences allouées aux composants",
    ),
    (
        "P5", "Acteurs vs systèmes", "Opérationnel",
        "stakeholders=[] malgré entités citées",
        "Distinguer acteurs de use case (stakeholders) vs frontières passives (external_systems)",
        "V2 : 5 acteurs vs 3 en V1",
    ),
    (
        "P6", "Typing des flux", "Fonctionnel",
        "0 flux typés sur 18",
        "Chaque flux doit avoir un flow_type (pneumatic, information, electric, thermal)",
        "Types de flux présents en V2",
    ),
    (
        "P7", "Composant::Fonction", "Fonctionnel",
        "Pas de lien composant-fonction",
        'Notation "NomComposant::NomFonction" pour les sous-fonctions',
        "3 notations composant::fonction en V2",
    ),
    (
        "P8", "Décomposition UC", "Opérationnel",
        "Use cases trop génériques",
        "Décomposer les UC par destination si plusieurs destinations mentionnées",
        "V2 : 8 UC vs 4",
    ),
]


# ═══════════════════════════════════════════════════════════════════════════════
# PARSING DES FICHIERS .md
# ═══════════════════════════════════════════════════════════════════════════════

def _first_code_block(text: str, lang: str = "") -> str:
    """Extrait le contenu du premier bloc de code dans *text*.
    Cherche d'abord ```<lang>, puis ``` générique."""
    if lang:
        m = re.search(rf"```{re.escape(lang)}\s*\n(.*?)```", text, re.DOTALL | re.IGNORECASE)
        if m:
            return m.group(1).strip()
    m = re.search(r"```(?:\w*)\s*\n(.*?)```", text, re.DOTALL)
    if m:
        return m.group(1).strip()
    # Aucun bloc de code : retourner le texte brut nettoyé
    return text.strip()


def parse_md_file(filepath: Path) -> dict:
    """
    Parse un fichier .md de résultats BAS et retourne un dict avec :
        prompt, json_response, sysml_code, warnings

    Supporte deux formats :
      • Format V2 : sections ## (Prompt / Réponse LLM (JSON) / Code SysML v2 / Warnings)
      • Format V1 : sections ### (Prompt envoyé au LLM / Réponse brute / Code SysML v2 généré)
    """
    out = {"prompt": "", "json_response": "", "sysml_code": "", "warnings": ""}

    if not filepath.exists():
        print(f"  ⚠️  Fichier introuvable — ignoré : {filepath}")
        return out

    content = filepath.read_text(encoding="utf-8")

    # ── Détection du format ──────────────────────────────────────────────────
    v2_format = bool(re.search(r"^## Prompt envoyé au LLM", content, re.MULTILINE))

    if v2_format:
        # Découper par "## " au début d'une ligne
        parts = re.split(r"^## ", content, flags=re.MULTILINE)
        for part in parts:
            key = part.split("\n", 1)[0].strip().lower()
            body = part[part.find("\n"):].strip() if "\n" in part else ""

            if key.startswith("prompt envoyé"):
                out["prompt"] = _first_code_block(body) or body
            elif key.startswith("réponse llm"):
                out["json_response"] = _first_code_block(body, "json") or body
            elif key.startswith("code sysml"):
                out["sysml_code"] = _first_code_block(body, "sysml") or body
            elif key.startswith("warnings"):
                # Supprimer les éventuels backtick vides
                warnings_text = re.sub(r"```\s*```", "", body).strip()
                out["warnings"] = warnings_text
    else:
        # Format V1 — sections ### ──────────────────────────────────────────
        parts = re.split(r"^### ", content, flags=re.MULTILINE)
        for part in parts:
            key = part.split("\n", 1)[0].strip().lower()
            body = part[part.find("\n"):].strip() if "\n" in part else ""

            if key.startswith("prompt envoyé"):
                raw = _first_code_block(body)
                # Tronquer les prompts très longs (RAG exemples inclus)
                if len(raw) > 4000:
                    raw = raw[:3980] + "\n\n[... tronqué — prompt RAG complet dans le fichier .md ...]"
                out["prompt"] = raw

                # Tenter d'extraire le JSON embarqué dans le prompt
                # (le JSON du modèle est passé comme contexte au LLM)
                json_m = re.search(r"```json\s*\n(.*?)```", body, re.DOTALL)
                if json_m:
                    out["json_response"] = json_m.group(1).strip()
                else:
                    # Chercher un grand bloc JSON {...}
                    json_m2 = re.search(r"(\{[\s\S]{200,}?\})\s*\n={3}", body)
                    if json_m2:
                        snippet = json_m2.group(1)
                        out["json_response"] = (
                            snippet[:2000] + "\n[... tronqué ...]" if len(snippet) > 2000
                            else snippet
                        )

            elif key.startswith("réponse brute"):
                out["sysml_code"] = _first_code_block(body, "sysml") or _first_code_block(body)

            elif key.startswith("code sysml v2 généré"):
                # Version "nettoyée" — on la préfère à la réponse brute
                sysml = _first_code_block(body, "sysml") or _first_code_block(body)
                if sysml:
                    out["sysml_code"] = sysml

    return out


# ═══════════════════════════════════════════════════════════════════════════════
# RÉCUPÉRATION DES ÉCHANGES  (API → fallback fichiers .md)
# ═══════════════════════════════════════════════════════════════════════════════

FILE_MAP = [
    ("V1", "Opérationnel", RESULTS_DIR / "style_formel"    / "operational.md"),
    ("V1", "Fonctionnel",  RESULTS_DIR / "style_formel"    / "functional.md"),
    ("V1", "Logique",      RESULTS_DIR / "style_formel"    / "logical.md"),
    ("V1", "Technique",    RESULTS_DIR / "style_formel"    / "technical.md"),
    ("V2", "Opérationnel", RESULTS_DIR / "style_formel_v2" / "operational.md"),
    ("V2", "Fonctionnel",  RESULTS_DIR / "style_formel_v2" / "functional.md"),
    ("V2", "Logique",      RESULTS_DIR / "style_formel_v2" / "logical.md"),
    ("V2", "Technique",    RESULTS_DIR / "style_formel_v2" / "technical.md"),
]


def _build_exchange(version: str, level: str, parsed: dict) -> dict:
    return {
        "system":        "BAS Silvercrest",
        "version":       version,
        "level":         level,
        "prompt":        parsed["prompt"],
        "json_response": parsed["json_response"],
        "sysml_code":    parsed["sysml_code"],
        "warnings":      parsed["warnings"],
        "comment":       COMMENTS.get((version, level), ""),
    }


def fetch_exchanges() -> list[dict]:
    """
    Tente de récupérer les échanges depuis l'API backend.
    Si l'API n'est pas joignable ou ne renvoie rien, lit les fichiers .md.
    """
    # ── Tentative API ────────────────────────────────────────────────────────
    try:
        import urllib.request
        import urllib.error

        api_reachable = False
        # Essayer plusieurs endpoints courants
        for health_path in ["/health", "/api/health", "/docs", "/"]:
            try:
                req = urllib.request.Request(
                    f"{API_BASE}{health_path}",
                    headers={"Accept": "application/json"},
                )
                resp = urllib.request.urlopen(req, timeout=3)
                # Tout code 2xx ou 3xx = serveur joignable
                api_reachable = resp.status < 400
                if api_reachable:
                    break
            except urllib.error.HTTPError as e:
                if e.code < 500:   # 4xx = serveur up, endpoint absent
                    api_reachable = True
                    break
            except Exception:
                pass  # Connexion refusée, timeout, etc.

        if api_reachable:
            print("✅  API disponible — récupération des sessions BAS...")
            return _fetch_from_api()
        raise ConnectionError("Aucun endpoint joignable")
    except Exception as exc:
        print(f"ℹ️  API non disponible ({exc.__class__.__name__}: {exc})")
        print("   → Fallback : lecture des fichiers .md")

    # ── Fallback fichiers .md ─────────────────────────────────────────────────
    return _fetch_from_files()


def _fetch_from_api() -> list[dict]:
    """Récupère les échanges BAS via l'API REST."""
    import urllib.request, urllib.error

    exchanges = []
    try:
        req = urllib.request.Request(
            f"{API_BASE}/api/sessions", headers={"Accept": "application/json"}
        )
        resp = urllib.request.urlopen(req, timeout=5)
        sessions = json.loads(resp.read().decode())

        bas_sessions = [
            s for s in sessions
            if any(kw in str(s).upper() for kw in ["BAS", "SILVERCREST", "BLEED"])
        ]
        print(f"   {len(bas_sessions)} session(s) BAS trouvée(s)")

        for session in bas_sessions:
            sid = session.get("id") or session.get("session_id") or session.get("uuid")
            if not sid:
                continue
            try:
                exch_req = urllib.request.Request(
                    f"{API_BASE}/api/v2/exchanges/{sid}",
                    headers={"Accept": "application/json"},
                )
                exch_resp = urllib.request.urlopen(exch_req, timeout=10)
                items = json.loads(exch_resp.read().decode())
                version = "V2" if "v2" in str(session.get("name", "")).lower() else "V1"
                for item in items:
                    level = item.get("level", "")
                    exchanges.append({
                        "system":        "BAS Silvercrest",
                        "version":       version,
                        "level":         level,
                        "prompt":        item.get("prompt", ""),
                        "json_response": item.get("json_response", ""),
                        "sysml_code":    item.get("sysml_code", ""),
                        "warnings":      "\n".join(item.get("warnings", [])),
                        "comment":       COMMENTS.get((version, level), ""),
                    })
            except Exception as e:
                print(f"  ⚠️  Erreur session {sid}: {e}")

    except Exception as e:
        print(f"⚠️  Erreur API sessions: {e}")

    if exchanges:
        print(f"✅  {len(exchanges)} échanges récupérés via API")
        return exchanges

    print("ℹ️  Aucun échange BAS via API — fallback fichiers .md")
    return _fetch_from_files()


def _fetch_from_files() -> list[dict]:
    """Lit les 8 fichiers .md et construit la liste d'échanges."""
    exchanges = []
    for version, level, filepath in FILE_MAP:
        rel = filepath.relative_to(BASE_DIR)
        print(f"  📄  {rel}  ({version} / {level})")
        parsed = parse_md_file(filepath)
        exchanges.append(_build_exchange(version, level, parsed))
    return exchanges


# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS MISE EN FORME EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def _apply_header(ws, n_cols: int):
    """Style la ligne 1 (en-têtes) d'une feuille."""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_WRAP_CTR


def _apply_row(ws, row_num: int, n_cols: int, *, alt: bool, tall: bool = False):
    """Style une ligne de données."""
    fill = ROW_FILL_ALT if alt else ROW_FILL_WHT
    align = ALIGN_WRAP_TOP if tall else ALIGN_WRAP_CTR
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.alignment = align
    if tall:
        ws.row_dimensions[row_num].height = 100


def _set_widths(ws, widths: list):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════════════════════════════
# CONSTRUCTION DES FEUILLES
# ═══════════════════════════════════════════════════════════════════════════════

def build_resume(wb: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet:
    ws = wb.create_sheet("Résumé")
    ws.append(SUMMARY_HEADERS)
    _apply_header(ws, len(SUMMARY_HEADERS))

    for i, row in enumerate(SUMMARY_ROWS, start=2):
        ws.append(list(row))
        _apply_row(ws, i, len(SUMMARY_HEADERS), alt=(i % 2 == 0))

    _set_widths(ws, [22, 8, 14, 30, 26, 16, 14, 14, 24])
    ws.freeze_panes = "A2"
    return ws


def build_exchanges(
    wb: openpyxl.Workbook, exchanges: list[dict]
) -> openpyxl.worksheet.worksheet.Worksheet:
    ws = wb.create_sheet("Échanges détaillés")
    ws.append(DETAIL_HEADERS)
    _apply_header(ws, len(DETAIL_HEADERS))

    for i, ex in enumerate(exchanges, start=1):
        row_num = i + 1
        ws.append([
            i,
            ex["system"],
            ex["version"],
            ex["level"],
            ex["prompt"],
            ex["json_response"],
            ex["sysml_code"],
            ex["comment"],
        ])
        _apply_row(ws, row_num, len(DETAIL_HEADERS), alt=(i % 2 == 0), tall=True)

    # Colonnes courtes pour ID/meta, larges pour le contenu textuel
    _set_widths(ws, [5, 20, 8, 14, 60, 60, 60, 40])
    ws.freeze_panes = "A2"
    return ws


def build_prompts(wb: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet:
    ws = wb.create_sheet("Prompts types")
    ws.append(PROMPTS_HEADERS)
    _apply_header(ws, len(PROMPTS_HEADERS))

    for i, row in enumerate(PROMPTS_TYPES, start=2):
        ws.append(list(row))
        _apply_row(ws, i, len(PROMPTS_HEADERS), alt=(i % 2 == 0))
        ws.row_dimensions[i].height = 60

    _set_widths(ws, [6, 22, 14, 42, 58, 36])
    ws.freeze_panes = "A2"
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# POINT D'ENTRÉE
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    print("\n🚀  Génération du livrable Excel — BAS Silvercrest échanges IA")
    print("=" * 62)

    # 1. Récupérer les échanges
    exchanges = fetch_exchanges()

    # 2. Créer le classeur
    wb = openpyxl.Workbook()
    # Supprimer la feuille vide créée par défaut
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    # 3. Construire les trois feuilles
    ws1 = build_resume(wb)
    ws2 = build_exchanges(wb, exchanges)
    ws3 = build_prompts(wb)

    # 4. Sauvegarder
    OUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_FILE)

    # 5. Rapport final
    size_ko = OUT_FILE.stat().st_size / 1024
    print("\n" + "=" * 62)
    print("✅  Fichier Excel généré avec succès !")
    print(f"   📊  Feuille 'Résumé'             : {ws1.max_row - 1} ligne(s) de données")
    print(f"   📊  Feuille 'Échanges détaillés' : {ws2.max_row - 1} ligne(s) de données")
    print(f"   📊  Feuille 'Prompts types'      : {ws3.max_row - 1} ligne(s) de données")
    print(f"   💾  Taille du fichier            : {size_ko:.1f} Ko")
    print(f"   📁  Chemin                       : {OUT_FILE.resolve()}")


if __name__ == "__main__":
    main()
